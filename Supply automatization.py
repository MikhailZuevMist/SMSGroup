import openpyxl
import os
import time


class Wb:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(filename=path)
        self.sheet = self.wb['Sheet1']


class Dicts:
    dictName = ['Наименование']
    dictAmount = ['Количество', 'Кол-во', 'К-во']
    dictMeasure = ['Ед.', 'Ед', 'Единицы', 'Ед.Изм']
    dictPrice = ['Цена']


def paste(data, sampleSheet, column):
    row = 20
    for i in data:
        if i is not None:
            s = i.split(' ')
            s.pop()
            i = ' '.join(s)
        if (column == 10 or column == 12 or column == 14) and i is not None:
            i = i.replace(',', '.')
            i = i.replace(' ', '')
            sampleSheet.cell(row=row, column=column).value = float(i) / 1.12
        else:
            sampleSheet.cell(row=row, column=column).value = i
        row += 1


def get_data(freerow, column, sheet):
    data = [sheet.cell(row=y, column=column).value for y in range(2, freerow + 1)]
    return data


def count_files():
    count = 0
    for path in os.scandir(r'C:\Users\ZuMi01\Documents\Power automate\DavidKP'):
        if path.is_file():
            count += 1
    return count


def copy_paste(freerow, i, sheet, sampleSheet, column):
    data = get_data(freerow, i, sheet)
    paste(data, sampleSheet, column)
    return data


def fill_sample(sheet, sampleSheet):
    freerow = sheet.max_row
    i = 0
    for cell in sheet[1]:
        i += 1
        if cell.value.split()[0] in Dicts.dictName:
            copy_paste(freerow, i, sheet, sampleSheet, 5)
        elif cell.value.split()[0] in Dicts.dictAmount:
            copy_paste(freerow, i, sheet, sampleSheet, 8)
        elif cell.value.split()[0] in Dicts.dictMeasure:
            copy_paste(freerow, i, sheet, sampleSheet, 9)


def fill_value(sheet1, saple_sheet, sheet2=None, sheet3=None):
    freerow = sheet1.max_row + 1
    i = 0
    if sheet2 == sheet3 is None:
        for cell in sheet1[1]:
            i += 1
            if cell.value.split()[0] in Dicts.dictPrice:
                copy_paste(freerow, i, sheet1, saple_sheet, 10)

    elif sheet2 != sheet3 is None:
        for cell in sheet1[1]:
            i += 1
            if cell.value.split()[0] in Dicts.dictPrice:
                data = copy_paste(freerow, i, sheet1, saple_sheet, 10)A

        i = 0
        for cell in sheet2[1]:
            i += 1
            if cell.value.split()[0] in Dicts.dictPrice:
                data = copy_paste(freerow, i, sheet2, saple_sheet, 12)
    else:
        for cell in sheet1[1]:
            i += 1
            if cell.value.split()[0] in Dicts.dictPrice:
                copy_paste(freerow, i, sheet1, saple_sheet, 10)
        i = 0
        for cell in sheet2[1]:
            i += 1
            if cell.value.split()[0] in Dicts.dictPrice:
                copy_paste(freerow, i, sheet2, saple_sheet, 12)
        i = 0
        for cell in sheet3[1]:
            i += 1
            if cell.value.split()[0] in Dicts.dictPrice:
                copy_paste(freerow, i, sheet2, saple_sheet, 14)


davidKP_folder = 'C:/Users/ZuMi01/Documents/Power automate/DavidKP'
samples_folder = 'C:/Users/ZuMi01/Documents/Power automate/Samples'
result_folder = 'C:/Users/ZuMi01/Documents/Power automate/Result'
file_number = 1
while True:
    while len(os.listdir(davidKP_folder)) == 0:
        pass
    time.sleep(1)
    count = count_files()
    if count == 1:
        sheet1 = Wb(f'{davidKP_folder}/1.xlsx')
        sample = Wb(f'{samples_folder}/Sample1.xlsx')
        fill_sample(sheet1.sheet, sample.sheet)
        fill_value(sheet1.sheet, sample.sheet)
        sample.wb.save(f'{result_folder}/{file_number}.xlsx')
    elif count == 2:
        sheet1 = Wb(f'{davidKP_folder}/1.xlsx')
        sheet2 = Wb(f'{davidKP_folder}/2.xlsx')
        sample = Wb(f'{samples_folder}/Sample2.xlsx')
        fill_sample(sheet1.sheet, sample.sheet)
        fill_value(sheet1.sheet, sample.sheet, sheet2.sheet)
        sample.wb.save(f'{result_folder}/{file_number}.xlsx')
    elif count == 3:
        sheet1 = Wb(f'{davidKP_folder}/1.xlsx')
        sheet2 = Wb(f'{davidKP_folder}/2.xlsx')
        sheet3 = Wb(f'{davidKP_folder}/3.xlsx')
        sample = Wb(f'{samples_folder}/Sample3.xlsx')
        fill_sample(sheet1.sheet, sample.sheet)
        fill_value(sheet1.sheet, sample.sheet, sheet2.sheet, sheet3.sheet)
        sample.wb.save(f'{result_folder}/{file_number}.xlsx')
    for f in os.listdir(davidKP_folder):
        os.remove(os.path.join(davidKP_folder, f))
    file_number += 1
    print(file_number)
